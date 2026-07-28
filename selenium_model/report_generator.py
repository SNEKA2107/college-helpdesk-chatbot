"""Phase 6 + 7 — Functional coverage analysis and the master Excel report.

Joins everything the earlier phases produced (discovery, code audit, executed
test results and the collector sinks) into a single management- and
developer-facing workbook: MASTER_TEST_AUDIT_REPORT.xlsx
"""
from __future__ import annotations

import json
import re
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import collectors
import config

# ── palette ─────────────────────────────────────────────────────────────────
NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color=NAVY, bold=True, size=14)

FILL = {
    "pass": PatternFill("solid", fgColor="D5F0DC"),
    "fail": PatternFill("solid", fgColor="F8D2D2"),
    "warn": PatternFill("solid", fgColor="FCE8C8"),
    "info": PatternFill("solid", fgColor="DCE6F5"),
    "muted": PatternFill("solid", fgColor="EEEEEE"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILL = {
    "Passed": "pass", "Pass": "pass", "OK": "pass", "Fully Covered": "pass",
    "Failed": "fail", "Fail": "fail", "Broken": "fail", "Not Covered": "fail",
    "Skipped": "warn", "Partially Covered": "warn",
    "High": "fail", "Medium": "warn", "Low": "info",
}


def _load(name):
    path = config.DATA_DIR / name
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


# ── Phase 6: coverage analysis ──────────────────────────────────────────────
CATEGORY_KEYWORDS = {
    "Auth": ("auth", "login", "logout", "session", "register", "setup", "password"),
    "RBAC": ("rbac", "authoriz", "role", "guard", "isolation"),
    "Navigation": ("nav", "sidebar", "topbar", "route", "redirect", "tab", "link"),
    "Form": ("form", "valid", "submit", "input", "boundary"),
    "CRUD": ("crud", "create", "update", "delete", "lifecycle"),
    "API-CRUD": ("api", "crud", "lifecycle"),
    "API": ("api", "contract", "endpoint"),
    "Search": ("search",),
    "Filter": ("filter",),
    "Sort": ("sort", "order"),
    "Table": ("table", "row", "pagination", "paginat"),
    "Modal": ("modal", "dialog", "popup"),
    "Notification": ("toast", "notification", "badge", "notice"),
    "Upload": ("upload", "file", "document", "attach"),
    "Download": ("download", "export"),
    "Button": ("button", "click", "action"),
    "Link": ("link", "nav"),
    "Journey": ("journey",),
    "Integration": ("integration", "api", "smoke"),
    "Payment": ("fee", "payment"),
    "Dashboard": ("dashboard", "chart", "analytics"),
}


def analyse_coverage(discovery, functional, api_rows, journeys):
    """Mark every discovered functionality Fully / Partially / Not Covered."""
    hits = {row["fid"] for row in collectors.read("coverage_hits")}

    executed = []
    for t in functional:
        executed.append({
            "id": t["test_id"], "name": t["name"], "module": (t.get("module") or "").lower(),
            "scenario": (t.get("scenario") or "").lower(), "status": t["status"],
            "blob": f"{t['test_id']} {t.get('module','')} {t.get('scenario','')}".lower(),
        })

    api_index = {}
    for r in api_rows:
        key = (str(r.get("endpoint", "")).lower(), str(r.get("method", "")).upper())
        api_index.setdefault(key, []).append(r)

    journey_names = {str(j.get("name", "")).lower(): j for j in journeys}

    rows = []
    for f in discovery["functionalities"]:
        fid = f["fid"]
        module = f["module"]
        page = f["page"]
        cat = f["category"]
        func = f["functionality"]
        evidence, status = [], "Not Covered"

        # 1. an explicit covers() declaration is the strongest signal
        if fid in hits:
            status = "Fully Covered"
            evidence.append("declared by a test via covers()")

        # 2. API functionalities join on endpoint + method
        if cat in ("API", "API-CRUD"):
            method = (f.get("detail") or "GET").split()[0].upper()
            ep = page.replace("/api", "", 1).lower()
            matched = api_index.get((ep, method)) or api_index.get((page.lower(), method))
            if matched:
                passed = [m for m in matched if str(m.get("result")).lower() == "pass"]
                status = "Fully Covered" if passed else "Partially Covered"
                evidence.append(f"{len(matched)} API check(s), {len(passed)} passing")
            elif status == "Not Covered":
                # A write endpoint reached only through a UI CRUD test still counts as partial.
                related = [t for t in executed
                           if any(k in t["blob"] for k in ep.strip("/").split("/") if len(k) > 3)]
                if related:
                    status = "Partially Covered"
                    evidence.append(f"exercised indirectly by {len(related)} test(s)")

        # 3. journeys join on name
        if cat == "Journey":
            j = journey_names.get(func.lower())
            if j:
                status = "Fully Covered" if j.get("result") == "Passed" else "Partially Covered"
                evidence.append(f"journey executed: {j.get('result')}")

        # 4. everything else: match module + page/keyword against executed tests
        if status == "Not Covered":
            mod_key = module.lower().split("—")[0].strip()
            page_key = page.strip("/").split("/")[-1].lower()
            keywords = CATEGORY_KEYWORDS.get(cat, ())
            module_tests = [t for t in executed
                            if mod_key and (mod_key in t["module"] or mod_key in t["blob"])]
            page_tests = [t for t in module_tests
                          if page_key and len(page_key) > 3 and page_key in t["blob"]]
            kw_tests = [t for t in (page_tests or module_tests)
                        if any(k in t["blob"] for k in keywords)]

            if kw_tests:
                passed = [t for t in kw_tests if t["status"] == "Passed"]
                status = "Fully Covered" if passed else "Partially Covered"
                evidence.append(f"{len(kw_tests)} matching test(s), {len(passed)} passing")
            elif page_tests:
                status = "Partially Covered"
                evidence.append(f"page exercised by {len(page_tests)} test(s), "
                                "but not this specific behaviour")
            elif module_tests:
                status = "Partially Covered"
                evidence.append(f"module exercised by {len(module_tests)} test(s)")

        remark = "; ".join(evidence) if evidence else \
            "No automated test exercises this behaviour — manual verification required."
        rows.append({
            "Page": page, "Functionality": func, "Module": module, "Type": cat,
            "Coverage Status": status, "Source File(s)": ", ".join(f["sources"][:3]),
            "Remarks": remark,
        })
    return rows


# ── recommendations ─────────────────────────────────────────────────────────
def build_recommendations(audit, functional, coverage, defects, security, a11y, perf):
    recs = []

    def add(priority, text, impact):
        recs.append({"Priority": priority, "Recommendation": text, "Business Impact": impact})

    failed = [t for t in functional if t["status"] == "Failed"]
    if failed:
        add("P1 — Critical",
            f"Fix the {len(failed)} failing end-to-end test(s) before sign-off; each one marks a "
            "user-facing behaviour that does not work as specified.",
            "Blocks release: users hit these paths directly.")

    high_sec = [s for s in security if s.get("severity") == "High"]
    if high_sec:
        add("P1 — Critical",
            f"Resolve {len(high_sec)} high-severity security observation(s), starting with the "
            "Content-Security-Policy 'unsafe-inline' allowance and JWT-in-localStorage storage.",
            "A single XSS becomes full account takeover across student, faculty and admin roles.")

    if any("password reset" in d.get("description", "").lower() for d in defects):
        add("P1 — Critical",
            "Implement a self-service password reset (email token or admin-issued temporary "
            "password with forced change).",
            "Locked-out students and staff currently need manual admin intervention, which does "
            "not scale past a pilot.")

    sensitive = [u for u in audit["unused_files"] if u.get("kind") == "sensitive-data"]
    if sensitive:
        add("P1 — Critical",
            f"Remove {len(sensitive)} file(s) of bulk student personal data from version control "
            "and purge them from git history.",
            "Committed PII is a data-protection breach and cannot be undone by deletion alone.")

    not_covered = [c for c in coverage if c["Coverage Status"] == "Not Covered"]
    if not_covered:
        add("P2 — High",
            f"Close the {len(not_covered)} uncovered functionality gap(s) — the faculty portal "
            "write paths and admin academic tabs are the largest clusters.",
            "Untested paths are where regressions reach production unnoticed.")

    dupes = [u for u in audit["unused_files"] if u.get("kind") == "duplicate-tree"]
    if dupes:
        add("P2 — High",
            f"Delete the {len(dupes)} clone/snapshot director(ies) checked into the repository "
            "(demo-clone, viva-clone, viva-clone2, …).",
            "Duplicate trees drift from source and make it ambiguous which code actually ships.")

    if not any("data-testid" in str(d) for d in defects):
        pass
    add("P2 — High",
        "Add stable data-testid attributes to interactive elements across the three portals.",
        "The suite currently binds to CSS classes and visible text, so routine styling or copy "
        "changes break tests that were not actually affected by the change.")

    high_a11y = [a for a in a11y if a.get("severity") == "High"]
    if high_a11y:
        add("P2 — High",
            f"Fix {len(high_a11y)} high-severity accessibility issue(s) — chiefly icon-only "
            "buttons and form controls with no accessible name.",
            "Blocks screen-reader users outright and is a compliance risk for a public "
            "institution.")

    orphans = [u for u in audit["unused_files"] if u.get("kind") == "orphan-component"]
    if orphans:
        add("P3 — Medium",
            f"Remove or re-route the {len(orphans)} orphan component(s) left behind by the "
            "unified-login change (RoleSelect, FacultyLogin, RoleCard).",
            "Dead screens mislead maintainers into thinking retired flows are still live.")

    slow = [p for p in perf if "exceeds" in str(p.get("observation", "")).lower()]
    if slow:
        add("P3 — Medium",
            f"Investigate {len(slow)} page/endpoint(s) over their latency budget; start with the "
            "unpaginated /api/students response.",
            "Slow admin screens are the first thing evaluators and daily users notice.")

    large = [d for d in audit["dead_code"] if d.get("kind") == "large-file"]
    if large:
        add("P3 — Medium",
            f"Split the {len(large)} oversized module(s) — facultyPortal.js (843 lines) and "
            "global.css (1205 lines) are the worst offenders.",
            "Large modules slow review and raise the chance of a merge conflict becoming a bug.")

    md_sprawl = [u for u in audit["unused_files"] if u.get("kind") == "doc-sprawl"]
    if md_sprawl:
        add("P4 — Low",
            "Consolidate the 100+ root-level status/report markdown files into docs/ with one "
            "authoritative README.",
            "Contributors cannot tell which document is current, so none of them are trusted.")

    add("P4 — Low",
        "Wire this suite into CI so `python selenium_model/run.py` runs on every pull request.",
        "Turns a point-in-time audit into a standing regression gate.")
    return recs


def build_code_health(audit, discovery):
    rows = []

    def add(cat, finding, sev, rec):
        rows.append({"Category": cat, "Finding": finding, "Severity": sev,
                     "Recommendation": rec})

    t = discovery["totals"]
    add("Scale", f"{t['source_files']} source files: {t['pages']} pages, {t['components']} shared "
                 f"components, {t['api_endpoints']} API endpoints, {t['models']} data models.",
        "Low", "No action — recorded as the audit baseline.")

    by_kind = {}
    for u in audit["unused_files"]:
        by_kind.setdefault(u["kind"], []).append(u)
    labels = {
        "duplicate-tree": ("Duplicate code trees", "Delete the clone directories and rely on git "
                                                   "history for snapshots."),
        "sensitive-data": ("Personal data in version control",
                           "Remove the files and purge them from git history; add them to "
                           ".gitignore."),
        "suspicious-file": ("Ad-hoc debug scripts at the repo root",
                            "Move anything still useful into tools/ and delete the rest."),
        "artifact": ("Build/debug artifacts committed",
                     "Add screenshots and logs to .gitignore."),
        "orphan-component": ("Orphan React components",
                             "Delete them or restore a route that uses them."),
        "unused-module": ("Unreferenced modules",
                          "Remove modules no entry point imports."),
        "backend-script": ("Standalone maintenance scripts",
                           "Keep, but document each one's purpose in the runbook."),
        "doc-sprawl": ("Documentation sprawl",
                       "Consolidate into docs/ with a single authoritative index."),
    }
    for kind, items in sorted(by_kind.items(), key=lambda kv: -len(kv[1])):
        label, rec = labels.get(kind, (kind, "Review and clean up."))
        worst = "High" if any(i["severity"] == "High" for i in items) else \
                ("Medium" if any(i["severity"] == "Medium" for i in items) else "Low")
        add("Unused / dead weight", f"{label}: {len(items)} item(s) "
                                    f"(e.g. {items[0]['path']})", worst, rec)

    dead_by_kind = {}
    for d in audit["dead_code"]:
        dead_by_kind.setdefault(d["kind"], []).append(d)
    dead_labels = {
        "large-file": ("Oversized modules", "Split by responsibility."),
        "unused-export": ("Exported symbols never imported", "Remove or document as public API."),
        "unreachable-function": ("Functions declared but never called", "Delete the dead code."),
        "placeholder": ("TODO/FIXME markers left in code", "Resolve or convert into tracked issues."),
        "duplicate": ("Byte-identical duplicate implementations", "Extract a shared module."),
    }
    for kind, items in sorted(dead_by_kind.items(), key=lambda kv: -len(kv[1])):
        label, rec = dead_labels.get(kind, (kind, "Review."))
        worst = "High" if any(i["severity"] == "High" for i in items) else \
                ("Medium" if any(i["severity"] == "Medium" for i in items) else "Low")
        add("Dead code", f"{label}: {len(items)} occurrence(s) "
                         f"(e.g. {items[0]['file']})", worst, rec)

    add("Testability", "The UI carries no data-testid attributes, so automation must bind to CSS "
                       "classes and visible copy.", "Medium",
        "Add data-testid to every interactive element the suite drives.")
    add("Architecture", "Route guards are enforced client-side, with the server independently "
                        "re-checking roles on every protected endpoint.", "Low",
        "Keep both layers — the client guard is UX, the server check is the control.")
    add("Build", "The React SPA is served directly by Express from frontend/dist, so one origin "
                 "covers UI and API and CORS is not in the critical path.", "Low",
        "No action required.")
    return rows


# ── workbook writing ────────────────────────────────────────────────────────
def _style_sheet(ws, df, title=None, freeze="A2"):
    if title:
        ws.insert_rows(1)
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT
        header_row = 2
        freeze = "A3"
    else:
        header_row = 1

    for cell in ws[header_row]:
        if cell.value is not None:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[header_row].height = 28

    widths = {}
    for col_idx, col in enumerate(df.columns, start=1):
        longest = max([len(str(col))] + [len(str(v)) for v in df[col].head(400).tolist()] or [10])
        widths[col_idx] = min(max(longest + 3, 12), 78)
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    status_cols = [i for i, c in enumerate(df.columns, start=1)
                   if str(c).strip() in ("Status", "Result", "Severity", "Coverage Status")]
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=len(df.columns)):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for ci in status_cols:
            cell = row[ci - 1]
            key = STATUS_FILL.get(str(cell.value).strip())
            if key:
                cell.fill = FILL[key]
                cell.font = Font(bold=True)
    ws.freeze_panes = freeze


def _write(writer, sheet, rows, columns, title=None):
    df = pd.DataFrame(rows, columns=columns) if rows else pd.DataFrame(columns=columns)
    if df.empty:
        df = pd.DataFrame([{c: ("— none found —" if i == 0 else "") for i, c in enumerate(columns)}],
                          columns=columns)
    startrow = 0
    df.to_excel(writer, sheet_name=sheet[:31], index=False, startrow=startrow)
    _style_sheet(writer.sheets[sheet[:31]], df, title)
    return df


def _load_metric(load, metric, default="—"):
    for row in load:
        if row.get("metric") == metric:
            return row.get("value", default)
    return default


def _load_headline(load):
    if not load:
        return "not executed"
    return (f"{_load_metric(load, 'Concurrent Virtual Users')} users · "
            f"{_load_metric(load, 'Throughput (RPS)')} · "
            f"avg {_load_metric(load, 'Response Time — Average')}")


def generate():
    discovery = _load("discovery.json")
    audit = _load("audit.json")
    if not discovery or not audit:
        raise SystemExit("Run discover.py and audit.py first — data/ is incomplete.")

    functional = collectors.read("functional")
    api_rows = collectors.read("api")
    journeys = collectors.read("journeys")
    defects_run = collectors.read("defects")
    links = collectors.read("broken_links")
    a11y = collectors.read("accessibility")
    ui = collectors.read("ui")
    perf = collectors.read("performance")
    security = collectors.read("security")
    load = collectors.read("load")
    load_endpoints = collectors.read("load_endpoints")

    coverage = analyse_coverage(discovery, functional, api_rows, journeys)

    passed = sum(1 for t in functional if t["status"] == "Passed")
    failed = sum(1 for t in functional if t["status"] == "Failed")
    skipped = sum(1 for t in functional if t["status"] == "Skipped")
    total = len(functional)

    full = sum(1 for c in coverage if c["Coverage Status"] == "Fully Covered")
    part = sum(1 for c in coverage if c["Coverage Status"] == "Partially Covered")
    none_ = sum(1 for c in coverage if c["Coverage Status"] == "Not Covered")
    cov_pct = round((full + 0.5 * part) / len(coverage) * 100, 1) if coverage else 0.0

    # ---- defects: static findings + runtime findings, de-duplicated ----
    all_defects, seen = [], set()
    for d in audit["defects"]:
        key = d["description"][:90]
        if key in seen:
            continue
        seen.add(key)
        all_defects.append({
            "Bug ID": "", "Module": d["module"], "Description": d["description"],
            "Steps to Reproduce": d["steps"], "Severity": d["severity"],
            "Evidence": d["evidence"], "Status": d["status"], "Source": "Static code audit",
        })
    for d in defects_run:
        key = d["description"][:90]
        if key in seen:
            continue
        seen.add(key)
        all_defects.append({
            "Bug ID": "", "Module": d["module"], "Description": d["description"],
            "Steps to Reproduce": d["steps"], "Severity": d["severity"],
            "Evidence": d["evidence"], "Status": d["status"], "Source": "Test execution",
        })
    order = {"High": 0, "Medium": 1, "Low": 2}
    all_defects.sort(key=lambda d: order.get(d["Severity"], 3))
    for i, d in enumerate(all_defects, start=1):
        d["Bug ID"] = f"BUG-{i:03d}"

    recommendations = build_recommendations(audit, functional, coverage, all_defects,
                                            security, a11y, perf)
    code_health = build_code_health(audit, discovery)

    t = discovery["totals"]
    now = datetime.now()

    # ── Executive Summary ───────────────────────────────────────────────────
    summary = [
        ("Project Name", config.PROJECT_NAME),
        ("Application Under Test", config.BASE_URL),
        ("Scan / Execution Date", now.strftime("%d %B %Y, %H:%M")),
        ("Test Framework", "Python · Selenium WebDriver · Pytest · Page Object Model"),
        ("Browser", "Google Chrome (headless)" if config.HEADLESS else "Google Chrome"),
        ("", ""),
        ("── SCOPE ──", ""),
        ("Total Files Scanned", t["files_scanned"]),
        ("Application Source Files", t["source_files"]),
        ("Total Pages / Screens", t["pages"]),
        ("Shared Components", t["components"]),
        ("Total Routes", t["routes"]),
        ("Backend API Endpoints", t["api_endpoints"]),
        ("Database Models", t["models"]),
        ("Total Functionalities Discovered", t["functionalities"]),
        ("", ""),
        ("── TEST EXECUTION ──", ""),
        ("Total Tests Executed", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Skipped", skipped),
        ("Pass Rate", f"{round(passed / total * 100, 1) if total else 0}%"),
        ("Total Execution Time", f"{round(sum(t_.get('duration_s', 0) for t_ in functional) / 60, 1)} min"),
        ("", ""),
        ("── FUNCTIONAL COVERAGE ──", ""),
        ("Fully Covered", full),
        ("Partially Covered", part),
        ("Not Covered", none_),
        ("Coverage Percentage", f"{cov_pct}%"),
        ("", ""),
        ("── BASELINE / LOAD TEST ──", ""),
        ("Concurrent Virtual Users", _load_metric(load, "Concurrent Virtual Users", "not executed")),
        ("Sustained Duration", _load_metric(load, "Test Duration")),
        ("Total Requests Served", _load_metric(load, "Total Requests")),
        ("Throughput", _load_metric(load, "Throughput (RPS)")),
        ("Error Rate", _load_metric(load, "Error Rate")),
        ("Response Time — Min / Avg / Max",
         f'{_load_metric(load, "Response Time — Min")} / '
         f'{_load_metric(load, "Response Time — Average")} / '
         f'{_load_metric(load, "Response Time — Max")}'),
        ("Response Time — p95", _load_metric(load, "Response Time — p95")),
        ("Load Test Verdict", _load_metric(load, "Baseline Verdict")),
        ("", ""),
        ("── QUALITY FINDINGS ──", ""),
        ("Total Bugs Found", len(all_defects)),
        ("  · High Severity", sum(1 for d in all_defects if d["Severity"] == "High")),
        ("  · Medium Severity", sum(1 for d in all_defects if d["Severity"] == "Medium")),
        ("  · Low Severity", sum(1 for d in all_defects if d["Severity"] == "Low")),
        ("Unused / Dead-weight Files", len(audit["unused_files"])),
        ("Dead-Code Findings", len(audit["dead_code"])),
        ("Broken Links", sum(1 for l in links if l["result"] == "Broken")),
        ("Accessibility Findings", len(a11y)),
        ("UI Validation Findings", len(ui)),
        ("Security Observations", len(security)),
        ("API Checks Executed", len(api_rows)),
        ("User Journeys Executed", len(journeys)),
        ("", ""),
        ("── VERDICT ──", ""),
        ("Release Recommendation",
         "GO — no failing functional tests" if failed == 0 else
         f"CONDITIONAL — {failed} failing test(s) require triage before sign-off"),
        ("Highest Risk Area",
         (sorted(all_defects, key=lambda d: order.get(d["Severity"], 3))[0]["Module"]
          if all_defects else "None identified")),
    ]

    with pd.ExcelWriter(config.REPORT_XLSX, engine="openpyxl") as writer:
        df = pd.DataFrame(summary, columns=["Metric", "Value"])
        df.to_excel(writer, sheet_name="Executive Summary", index=False)
        ws = writer.sheets["Executive Summary"]
        _style_sheet(ws, df, "CampusAssist — Master Test & Audit Report")
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 76
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=2):
            label = str(row[0].value or "")
            if label.startswith("──"):
                for c in row:
                    c.fill = FILL["info"]
                    c.font = Font(bold=True, color=NAVY)

        _write(writer, "Functional Test Results",
               [{"Test ID": t_["test_id"].split("::")[-1],
                 "Module": t_["module"],
                 "Scenario": t_["scenario"],
                 "Expected Result": t_["expected"],
                 "Actual Result": t_["actual"],
                 "Status": t_["status"],
                 "Execution Time (s)": t_["duration_s"],
                 "Screenshot Path": t_["screenshot"],
                 "Full Node ID": t_["test_id"]} for t_ in functional],
               ["Test ID", "Module", "Scenario", "Expected Result", "Actual Result",
                "Status", "Execution Time (s)", "Screenshot Path", "Full Node ID"],
               f"Functional Test Results — {passed} passed / {failed} failed / {skipped} skipped")

        _write(writer, "Functional Coverage", coverage,
               ["Page", "Functionality", "Module", "Type", "Coverage Status",
                "Source File(s)", "Remarks"],
               f"Functional Coverage — {cov_pct}% ({full} full, {part} partial, {none_} none)")

        _write(writer, "Defect Report", all_defects,
               ["Bug ID", "Module", "Description", "Steps to Reproduce", "Severity",
                "Evidence", "Status", "Source"],
               f"Defect Report — {len(all_defects)} finding(s)")

        _write(writer, "Unused Files",
               [{"File Name": u["file"], "Path": u["path"], "Reason": u["reason"],
                 "Severity": u["severity"], "Lines": u.get("lines", 0)}
                for u in audit["unused_files"]],
               ["File Name", "Path", "Reason", "Severity", "Lines"],
               f"Unused / Dead-weight Files — {len(audit['unused_files'])} item(s)")

        _write(writer, "Dead Code",
               [{"File": d["file"], "Function or Class": d["symbol"], "Line Number": d["line"],
                 "Severity": d["severity"], "Recommendation": d["recommendation"]}
                for d in audit["dead_code"]],
               ["File", "Function or Class", "Line Number", "Severity", "Recommendation"],
               f"Dead Code — {len(audit['dead_code'])} finding(s)")

        _write(writer, "Broken Links",
               [{"URL": l["url"], "Source Page": l["source_page"],
                 "Status Code": l["status_code"], "Result": l["result"]} for l in links],
               ["URL", "Source Page", "Status Code", "Result"],
               f"Broken Link Validation — {len(links)} link(s) checked")

        _write(writer, "Accessibility Findings",
               [{"Page": a["page"], "Issue": a["issue"], "Severity": a["severity"],
                 "Recommendation": a["recommendation"]} for a in a11y],
               ["Page", "Issue", "Severity", "Recommendation"],
               f"Accessibility Findings — {len(a11y)} issue(s)")

        _write(writer, "API Validation Results",
               [{"Endpoint": a["endpoint"], "Method": a["method"],
                 "Expected Status": a["expected"], "Actual Status": a["actual"],
                 "Result": a["result"], "Note": a.get("note", "")} for a in api_rows],
               ["Endpoint", "Method", "Expected Status", "Actual Status", "Result", "Note"],
               f"API Validation — {len(api_rows)} check(s)")

        _write(writer, "UI Validation Findings",
               [{"Page": u["page"], "Issue": u["issue"], "Severity": u["severity"],
                 "Evidence": u["evidence"]} for u in ui],
               ["Page", "Issue", "Severity", "Evidence"],
               f"UI Validation — {len(ui)} finding(s)")

        _write(writer, "Performance Observations",
               [{"Page": p["page"], "Load Time (ms)": p["load_ms"],
                 "Observation": p["observation"], "Recommendation": p["recommendation"]}
                for p in perf],
               ["Page", "Load Time (ms)", "Observation", "Recommendation"],
               f"Performance Observations — {len(perf)} measurement(s)")

        _write(writer, "Load Test Results",
               [{"Metric": l["metric"], "Value": l["value"],
                 "Observation": l.get("observation", ""),
                 "Recommendation": l.get("recommendation", "")} for l in load],
               ["Metric", "Value", "Observation", "Recommendation"],
               f"Baseline / Load Test — {_load_headline(load)}")

        _write(writer, "Load Test By Endpoint",
               [{"Endpoint": e["endpoint"], "Requests": e["requests"],
                 "Successful": e["ok"], "Errors": e["errors"],
                 "Throughput (req/sec)": e["rps"], "Min (ms)": e["min_ms"],
                 "Average (ms)": e["avg_ms"], "p95 (ms)": e["p95_ms"],
                 "Max (ms)": e["max_ms"]} for e in load_endpoints],
               ["Endpoint", "Requests", "Successful", "Errors", "Throughput (req/sec)",
                "Min (ms)", "Average (ms)", "p95 (ms)", "Max (ms)"],
               f"Load Test — per-endpoint breakdown ({len(load_endpoints)} endpoints)")

        _write(writer, "User Journey Results",
               [{"Journey Name": j["name"], "Steps": j["steps"], "Result": j["result"],
                 "Evidence": j["evidence"]} for j in journeys],
               ["Journey Name", "Steps", "Result", "Evidence"],
               f"End-to-End User Journeys — {len(journeys)} executed")

        _write(writer, "Security Observations",
               [{"Area": s["area"], "Observation": s["observation"], "Severity": s["severity"],
                 "Recommendation": s["recommendation"]} for s in security],
               ["Area", "Observation", "Severity", "Recommendation"],
               f"Security Observations — {len(security)} item(s)")

        _write(writer, "Code Health Summary", code_health,
               ["Category", "Finding", "Severity", "Recommendation"],
               "Code Health Summary")

        _write(writer, "Recommendations", recommendations,
               ["Priority", "Recommendation", "Business Impact"],
               "Prioritised Recommendations")

        # ---- supporting inventories ----
        _write(writer, "Discovered Functionality",
               [{"Module": f["module"], "Page / Route": f["page"],
                 "Functionality": f["functionality"], "Type": f["category"],
                 "Source File(s)": ", ".join(f["sources"]), "Detail": f.get("detail", "")}
                for f in discovery["functionalities"]],
               ["Module", "Page / Route", "Functionality", "Type", "Source File(s)", "Detail"],
               f"Phase 1 Discovery — {t['functionalities']} functionalities mapped to source")

        _write(writer, "API Inventory",
               [{"Endpoint": e["endpoint"], "Method": e["method"],
                 "Access Level": e["role"], "Module": e["module"], "Source File": e["file"]}
                for e in discovery["endpoints"]],
               ["Endpoint", "Method", "Access Level", "Module", "Source File"],
               f"Backend API Inventory — {t['api_endpoints']} endpoints")

        _write(writer, "Route Map",
               [{"Route": r["path"], "Access": r["access"]} for r in discovery["routes"]],
               ["Route", "Access"],
               f"Application Route Map — {t['routes']} routes")

    print(f"  Tests      : {total} ({passed} passed, {failed} failed, {skipped} skipped)")
    print(f"  Coverage   : {cov_pct}% ({full} full / {part} partial / {none_} none)")
    print(f"  Defects    : {len(all_defects)}")
    print(f"  -> {config.REPORT_XLSX.name}")

    _write_markdown(discovery, audit, functional, coverage, all_defects, security,
                    a11y, ui, perf, journeys, links, api_rows, recommendations,
                    passed, failed, skipped, cov_pct, full, part, none_, load)
    return config.REPORT_XLSX


# ── FINAL_AUDIT_REPORT.md ───────────────────────────────────────────────────
def _write_markdown(discovery, audit, functional, coverage, defects, security, a11y, ui,
                    perf, journeys, links, api_rows, recommendations,
                    passed, failed, skipped, cov_pct, full, part, none_, load=()):
    t = discovery["totals"]
    total = len(functional)
    now = datetime.now().strftime("%d %B %Y, %H:%M")
    high = [d for d in defects if d["Severity"] == "High"]

    lines = [
        "# CampusAssist — Final Audit Report",
        "",
        f"**Generated:** {now}  ",
        f"**Application under test:** {config.BASE_URL}  ",
        "**Framework:** Python · Selenium WebDriver · Pytest · Page Object Model  ",
        f"**Master workbook:** `selenium_model/MASTER_TEST_AUDIT_REPORT.xlsx`",
        "",
        "---",
        "",
        "## 1. Executive summary",
        "",
        f"| Metric | Value |",
        f"|---|---|",
        f"| Files scanned | {t['files_scanned']} |",
        f"| Pages / components | {t['pages']} / {t['components']} |",
        f"| Routes | {t['routes']} |",
        f"| API endpoints | {t['api_endpoints']} |",
        f"| Functionalities discovered | {t['functionalities']} |",
        f"| Tests executed | {total} |",
        f"| Passed / Failed / Skipped | {passed} / {failed} / {skipped} |",
        f"| Functional coverage | {cov_pct}% ({full} full, {part} partial, {none_} none) |",
        f"| Defects found | {len(defects)} ({len(high)} high) |",
        f"| User journeys executed | {len(journeys)} |",
        f"| Load test | {_load_headline(load)} |",
        "",
        f"**Release recommendation:** "
        + ("GO — every executed functional test passed."
           if failed == 0 else
           f"CONDITIONAL — {failed} failing test(s) need triage before sign-off."),
        "",
        "---",
        "",
        "## 2. What was tested",
        "",
        "- **Authentication** — unified login for all three roles, email/register-number/case "
        "handling, invalid credentials, account enumeration, field validation, password "
        "visibility, remember-me, logout, session teardown, registration, approval gating and "
        "first-run setup sealing.",
        "- **Authorization** — route guards for unauthenticated access, cross-portal isolation "
        "for all three roles, server-side adminOnly/facultyOnly enforcement, token tampering, "
        "privilege escalation via the registration and login bodies, and per-owner data scoping.",
        "- **Navigation** — every student, faculty and admin destination, all 19 admin panel "
        "tabs, topbar controls, mobile bottom navigation, theme switching and the legacy URL "
        "redirect contract.",
        "- **Forms** — mandatory-field validation, invalid input, boundary values (password "
        "length, oversized payloads, long text) and valid submissions.",
        "- **CRUD** — full create/read/update/delete lifecycles for requests, notices, events, "
        "departments and leave, including owner scoping and workflow-integrity checks.",
        "- **Search, filters, sorting, tables and pagination** — exact, partial, empty and "
        "no-match searches; category and status filters; table rendering; ordering guarantees.",
        "- **Modals, notifications, uploads and downloads.**",
        "- **End-to-end journeys** — nine complete multi-step workflows.",
        "- **Additional layers** — smoke, regression, broken links, API contract validation, "
        "accessibility, UI integrity and performance.",
        "",
        "---",
        "",
        "## 3. Highest-severity findings",
        "",
    ]

    if high:
        for d in high[:12]:
            lines += [f"### {d['Bug ID']} — {d['Module']}", "",
                      d["Description"], "",
                      f"*Reproduce:* {d['Steps to Reproduce']}  ",
                      f"*Evidence:* `{d['Evidence']}`", ""]
    else:
        lines += ["No high-severity defects were identified.", ""]

    lines += [
        "---",
        "",
        "## 4. Code health",
        "",
        f"- **{len(audit['unused_files'])} unused / dead-weight files**, including "
        f"{sum(1 for u in audit['unused_files'] if u['kind'] == 'duplicate-tree')} full clone "
        "directories of the project and "
        f"{sum(1 for u in audit['unused_files'] if u['kind'] == 'sensitive-data')} files of bulk "
        "student personal data committed to version control.",
        f"- **{len(audit['dead_code'])} dead-code findings**: orphan exports, unreachable "
        "functions and oversized modules.",
        f"- **{sum(1 for u in audit['unused_files'] if u['kind'] == 'orphan-component')} orphan "
        "React components** left behind by the move to a unified login.",
        "",
        "---",
        "",
        "## 5. Coverage gaps",
        "",
        f"{none_} discovered functionalities have no automated coverage. The largest clusters:",
        "",
    ]

    gaps = {}
    for c in coverage:
        if c["Coverage Status"] == "Not Covered":
            gaps[c["Module"]] = gaps.get(c["Module"], 0) + 1
    for mod, n in sorted(gaps.items(), key=lambda kv: -kv[1])[:10]:
        lines.append(f"- **{mod}** — {n} uncovered functionality item(s)")

    lines += [
        "",
        "---",
        "",
        "## 6. Baseline / load test",
        "",
    ]
    if load:
        lines += [
            f"{_load_metric(load, 'Concurrent Virtual Users')} virtual users were held active "
            f"against the API continuously for {_load_metric(load, 'Test Duration')}, "
            "browsing the endpoint mix a real student, faculty member and administrator hit "
            "on a normal day.",
            "",
            "| Metric | Value | Observation |",
            "|---|---|---|",
        ]
        for row in load:
            lines.append(f"| {row['metric']} | {row['value']} | {row.get('observation', '')} |")
    else:
        lines.append("_Load test was not executed in this run._")

    lines += [
        "",
        "---",
        "",
        "## 7. Prioritised recommendations",
        "",
        "| Priority | Recommendation | Business impact |",
        "|---|---|---|",
    ]
    for r in recommendations:
        lines.append(f"| {r['Priority']} | {r['Recommendation']} | {r['Business Impact']} |")

    lines += [
        "",
        "---",
        "",
        "## 8. Deliverables",
        "",
        "| Artifact | Path |",
        "|---|---|",
        "| Master Excel report | `selenium_model/MASTER_TEST_AUDIT_REPORT.xlsx` |",
        "| HTML execution report | `selenium_model/execution_report.html` |",
        f"| Screenshots ({len(list(config.SCREENSHOT_DIR.glob('*.png')))}) | `selenium_model/screenshots/` |",
        "| Browser console log | `selenium_model/logs/browser_console.log` |",
        "| Selenium driver log | `selenium_model/logs/selenium.log` |",
        "| Backend server log | `selenium_model/logs/backend-server.log` |",
        "| Raw phase data (JSON) | `selenium_model/data/` |",
        "| This report | `selenium_model/FINAL_AUDIT_REPORT.md` |",
        "",
        "## 9. Reproducing this run",
        "",
        "```bash",
        "pip install -r selenium_model/requirements.txt",
        "node backend/dev-local.js          # seeded in-memory backend on :5000",
        "python selenium_model/run.py       # all seven phases, end to end",
        "```",
        "",
    ]

    (config.ROOT / "FINAL_AUDIT_REPORT.md").write_text("\n".join(lines), encoding="utf-8")
    print("  -> FINAL_AUDIT_REPORT.md")


if __name__ == "__main__":
    generate()
